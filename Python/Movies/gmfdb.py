# Alex Miller
# scrape for movies

import random, sys, requests, bs4
import openpyxl as xl
from openpyxl.styles import Font
import re
import shutil
from pathlib import Path

import smtplib
from email.message import EmailMessage

def main():
    # https://moviesjoy.is/filter?type=movie&quality=all&release_year=all&genre=all&country=all
    # https://moviesjoy.is/filter?type=movie&quality=all&release_year=all&genre=all&country=all&page=2
    '''
    flw-item
    film-name
    fdi-item
    fdi-duration

    '''
    try:
        website = 'https://moviesjoy.is/movie'
        getStuff(website, 2, 2, 1)
    except:
        #SendEmail('9062354870@vtext.com', 'Error', 'Error in main')
        print('#'*50)
        print('ERROR IN MAIN')
        sys.exit()
    '''
    website = 'https://moviesjoy.is/movie?page=119'
    getStuff(website, 2, 2, 1)
    '''

def getStuff(website, page, start, ct): # method for actually getting the movies
    res = requests.get(website)
    content = res.content
    try:
        res.raise_for_status()
    except Exception as exc:
        print('There was a problem with {}'.format(exc))
        #SendEmail('9062354870@vtext.com', 'Error', 'There was an error, come quick')
    
    soup = bs4.BeautifulSoup(content, 'html.parser')

    wb = xl.load_workbook('moviedb.xlsx') # excel sheet
    sheet = wb['Movies']

    elems = soup.find_all('h2', class_='film-name') # film names
    
    # fdi-item fdi-duration && fdi-item
    # 1239 pages total
    tyelems = soup.select('div.fd-infor span.fdi-item')
    texts = [element.text for element in tyelems]

    font = Font(name='Times New Roman', size=10,) # changing the font of the excel file
    hreftags = soup.select('h2 a')

    links = [hreftags.get('href') for hreftags in hreftags] # getting the links

    t = 0
    p = 0
    begin = start
    for i in range(2, len(elems)+2): # putting each movie in the database
        cell = 'A' + str(start)
        sheet[cell] = elems[p].getText()
        sheet[cell].font = font
        print('Added #{}: '.format(ct) + elems[p].getText()) # text to know that its working
        ct += 1

        cell = 'E' + str(start)
        if not links[p].startswith('https://'): # bad link
            link = 'https://moviesjoy.is' + links[p]
            sheet[cell] = link
        else:
            sheet[cell] = links[p]
        
        genre = getGenre(links[p], wb, sheet) # genres
        cell = 'C' + str(start)
        sheet[cell] = genre
        sheet[cell].font = font
        
        cell = 'B' + str(start) # years
        if texts[t] == 'N/A':
            sheet[cell] = 'N/A'
        else:
            sheet[cell] = texts[t]
        sheet[cell].font = font

        t += 1
        cell = 'D' + str(start) # length
        l = texts[t]
        if l == 'N/A':
            sheet[cell] = 'N/A'
        else:
            b = l[:len(l)-1] # without 'm'
            mtoh = MtoH(b)
            length = '{}'.format(mtoh)

            sheet[cell] = length
        sheet[cell].font = font

        t += 1
        p += 1
        start += 1

        wb.save('moviedb.xlsx') # save the file
        path = Path.cwd()
        shutil.copy(path / 'moviedb.xlsx', path / 'moviedb-copy.xlsx')

    for i in range(2, sheet.max_row+1):
        sheet.row_dimensions[i].height = 40
    wb.save('moviedb.xlsx')

    # double check everything was added right
    for i in range(begin,sheet.max_row+1):
        cellvalue = sheet['B' + str(i)].value
        if not cellvalue.isdecimal() or cellvalue != 'N/A':
            #SendEmail('9062354870@vtext.com', 'Error', 'Year is wrong')
            print('#'*50)
            print('YEAR IS WRONG')
            sys.exit()
        cellvalue = sheet['D' + str(i)].value
        if CheckType(cellvalue) == None:
            if cellvalue != 'N/A':
                print('#'*50)
                print('TIME IS WRONG')
                #SendEmail('9062354870@vtext.com', 'Error', 'Time is wrong')
                sys.exit()

    # go to next page
    website = 'https://moviesjoy.is/movie' + '?page={}'.format(page)
    print('''
========

   {}

========'''.format(page))
    page += 1
    if page == 1242:
        print('#'*50)
        print('IT IS DONE')
        #SendEmail('9062354870@vtext.com', 'Done', 'The program has reached it\'s end')
        sys.exit()
    #if page % 50 == 0:
        #SendEmail('9062354870@vtext.com', 'Update', 'We have hit {} pages, and {} movies'.format(page, ct))
    getStuff(website, page, start, ct)

def MtoH(l):
    h = 0
    while int(l) >= 60:
        l = int(l) - 60
        h += 1
    if h > 0:
        if l == 0:
            t = '{} H'.format(h)
        else:
            t = '{} H {} m'.format(h, l)
    else:
        t = '{} m'.format(l)
    return t

def CheckType(cell):
    HMMFormat = re.compile(r'\d H \d\d m')
    HMFormat = re.compile(r'\d H \d m')
    HFormat = re.compile(r'\d H')
    MMFormat = re.compile(r'\d\d m')
    MFormat = re.compile(r'\d m')
    mo = HMMFormat.search(tname)
    if mo == None:
        mo = HMFormat.search(tname)
        if mo == None:
            mo = HFormat.search(tname)
            if mo == None:
                mo = MMFormat.search(tname)
                if mo == None:
                    mo = MFormat.search(tname)
                    if mo == None: # if user typed in a wrong format
                        return None
    if len(tname) == 8:
        hm = HMMFormat
    elif len(tname) == 7:
        hm = HMFormat
    elif len(tname) == 3:
        if tname.endswith('H'):
            hm = HFormat
        elif tname.endswith('m'):
            hm = MFormat
    elif len(tname) == 4:
        hm = MMFormat
    
    return hm

def getGenre(link, wb, sheet): # getting the genres of the movies
    if not link.startswith('https://'): # if not a good link
        link = 'https://moviesjoy.is' + link
    rs = requests.get(link)
    content = rs.content
    try:
        rs.raise_for_status()
    except Exception as exc:
        print('There was a problem with {}'.format(exc))
        #SendEmail('9062354870@vtext.com', 'Error', 'There was an error, come quick')

    gsoup = bs4.BeautifulSoup(content, 'html.parser')
    p = gsoup.select('div.row-line a')
    if p == []:
        return 'N/A'
    genre = p[0].getText()

    return genre

def SendEmail(to, subject, body):
    msg = EmailMessage()
    msg.set_content(body)
    msg['subject'] = subject
    msg['to'] = to

    user = 'pythonerrors404@gmail.com'
    msg['from'] = user
    password = 'xezzksakyhqxmsaa'

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(user, password)
    server.send_message(msg)

    server.quit()

if __name__ == '__main__': # run main
    #SendEmail('9062354870@vtext.com', 'Start', 'Program has started')
    main()
