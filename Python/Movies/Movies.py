# Alex Miller
# Accessing the movie database excel file

import random, sys, os
import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import pyperclip

def main(): # main method
    print()
    # ask for a command
    PrintStyle('What would you like to do? ')
    cmd = input('''
    
    (s)     Search                (g)     List of all genres

    (r)     Random Movie          (y)     List of all years

    (cls)   Clear Screen          (l)     List of all lengths

    (q)     Quit

> ''')
    processcmd(cmd)

def processcmd(cmd): # does stuff with the given command
    srcfile = 'moviedb.xlsx'
    wb = xl.load_workbook(srcfile)
    sheet = wb['Movies']
    if cmd.lower().startswith('s'): # searching
        ClearSheet(wb)
        SearchType(wb, sheet)
    elif cmd.lower().startswith('g'):
        PrintTypes(wb, sheet, 'C')
    elif cmd.lower().startswith('y'):
        PrintTypes(wb, sheet, 'B')
    elif cmd.lower().startswith('l'):
        PrintTypes(wb, sheet, 'D')
    elif cmd.lower().startswith('r'): # random movie
        Random(wb, sheet)
    elif cmd.lower().startswith('q'): # quit
        print()
        PrintStyle('Thanks!')
        print()
        sys.exit()
    elif cmd.lower() == 'cls': # clear screen
        for i in range(50):
            print()
        main()
    else: # not a command
        print()
        PrintErrors('That didn\'t work, Try again ')
        main() # ask again


def SearchType(wb, sheet): # search for movies by a certain parameter
    print()
    PrintStyle('How would you like to search? ')
    print()
    t = input('''

    (n)    Name                   (l)    Length

    (y)    Year                   (g)    Genre

    (m)    Main Menu              (a)    More than one of above
            
>''')
    print()
    c = ''
    if t.lower().startswith('y'): # search by year
        c = 'B'
        PrintStyle('Enter the year you want to search (can use <, >, <=, >= before number Ex: \'<= 2023\') ')
        print()
        tname = input('> ')
    elif t.lower().startswith('n'):
        Search(wb, sheet)
    elif t.lower().startswith('g'): # search by genre
        c = 'C'
        PrintStyle('Enter the genre you want to search ')
        print()
        tname = input('> ')
    elif t.lower().startswith('l'): # search by length
        c = 'D'
        PrintStyle('Enter the length you want to search (Ex: 1 H 10 m, 1 H 5 m, 10 m, 5 m, 2 H) OR (<= 1 H, > 2H 30m) etc')
        print()
        tname = input('> ')
        if 'H' not in tname and ' m' not in tname: # wrong format
            PrintErrors('Wrong format, Try again')
            SearchType(wb, sheet)
    elif t.lower().startswith('a'):
        MoreFilters(wb, sheet)
    elif t.lower().startswith('m'): # back to main menu
        main()
    else: # not a command
        print()
        PrintErrors('That didn\'t work, Try again')
        SearchType(wb, sheet) # ask again
    ct = 0
    for i in range(2, sheet.max_row+1): # searches for the movies
        cell = c + str(i)
        if str(sheet[cell].value) == 'N/A' or not (sheet[cell].value).isdecimal():
            continue
        if c == 'B' or c == 'D': # length or year
            if tname.lower().startswith('> '): # greater than
                if Greater(wb, sheet, tname, cell, i, ct) == True:
                    PrintMovie(wb, sheet, i) # prints the movie
                    ImportMovie(wb, sheet, i, getindex(wb))
                    ct += 1
            elif tname.lower().startswith('< '): # less than
                if Less(wb, sheet, tname, cell, i, ct) == True:
                    PrintMovie(wb, sheet, i) # prints the movie
                    ImportMovie(wb, sheet, i, getindex(wb))
                    ct += 1
            elif tname.lower().startswith('>= '): # greater or equal
                if GreaterEqual(wb, sheet, tname, cell, i, ct) == True:
                    PrintMovie(wb, sheet, i) # prints the movie
                    ImportMovie(wb, sheet, i, getindex(wb))
                    ct += 1
            elif tname.lower().startswith('<= '): # less or equal
                if LessEqual(wb, sheet, tname, cell, i, ct) == True:
                    PrintMovie(wb, sheet, i) # prints the movie
                    ImportMovie(wb, sheet, i, getindex(wb))
                    ct += 1
            else: # if there are no expression symbols
                if c == 'D':
                    hm = getType(wb, sheet, tname)
                    if hm == None:
                        PrintErrors('Wrong format, Try again')
                        SearType(wb, sheet)
                else:
                    hm = 0
                if Equal(wb, sheet, tname, cell, i, ct, hm) == True:
                    PrintMovie(wb, sheet, i) # prints the movie
                    ImportMovie(wb, sheet, i, getindex(wb))
                    ct += 1
        else: # genre
            if Equal(wb, sheet, tname, cell, i, ct, 0) == True:
                PrintMovie(wb, sheet, i) # prints the movie
                ImportMovie(wb, sheet, i, getindex(wb))
                ct += 1
    if ct == 0: # if no movies with parameters are found
        PrintErrors('There was no movies that includes: ' + tname)
    else:
        print()
        PrintStyle('All movies with requested parameters are in the database under \'ParMovies\' ' + '\n' + 'Would you like to look at them? (y/n)')
        print()
        wb.save('moviedb.xlsx')
        look = input('> ')
        if look.lower().startswith('y'):
            try:
                os.system('start SCALC.EXE moviedb.xlsx')
            except:
                try:
                    os.system('start EXCEL.EXE moviedb.clsx')
                except:
                    PrintError('Cannot open up the file')
        else:
            main()
    main()

def MoreFilters(wb, sheet):
    print()
    PrintStyle('Enter the year (n if none) ')
    print()
    year = input('> ')
    if year.startswith('<= '):
        Ysymbol = '<='
    elif year.startswith('< '):
        Ysymbol = '<'
    elif year.startswith('>='):
        Ysymbol = '>='
    elif year.startswith('> '):
        Ysymbol = '>'
    else:
        if not year.isdecimal() and year != 'n':
            PrintErrors('Year must be a number, Try again')
            MoreFilters(wb, sheet)
        Ysymbol = ''
    print()
    PrintStyle('Enter the length (n if none) ')
    print()
    length = input('> ')
    if length != 'n':
        if length.lower().startswith('<= '): # less equal
            symbol = '<='
        elif length.lower().startswith('< '): # less
            symbol = '<'
        elif length.lower().startswith('>= '): # great equal
            symbol = '>='
        elif length.lower().startswith('> '): # great
            symbol = '>'
        else:
            hm = getType(wb, sheet, length)
            if hm == None:
                PrintErrors('Wrong format, Try again')
                MoreFilters(wb, sheet)
            if not hm.search(length):
                PrintErrors('Wrong format, Try again (ex: 1 H 10 m) ')
                MoreFilters(wb, sheet)
            symbol = ''

    print()
    PrintStyle('Enter the genre (n if none) ')
    print()
    genre = input('> ')
    print()
    print(year, length, genre)
    ct = 0
    if year == 'n' and length != 'n' and genre != 'n':
        for i in range(2, sheet.max_row+1):
            lcell = 'D' + str(i)
            gcell = 'C' + str(i)
            if str(sheet[gcell].value).lower() == genre.lower():
                if symbol == '>':
                    if Greater(wb, sheet, length, lcell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                elif symbol == '>=':
                    if GreaterEqual(wb, sheet, length, lcell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                elif symbol == '<':
                    if Less(wb, sheet, length, lcell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                elif symbol == '<=':
                    if LessEqual(wb, sheet, length, lcell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                else:
                    if str(sheet[lcell].value).lower() == length.lower():
                        PrintMovie(wb, sheet, i)
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
    elif year != 'n' and length == 'n' and genre != 'n':
        for i in range(2, sheet.max_row+1):
            ycell = 'B' + str(i)
            gcell = 'C' + str(i)
            if str(sheet[gcell].value).lower() == genre.lower():
                if Ysymbol == '>':
                    if Greater(wb, sheet, year, ycell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                elif Ysymbol == '>=':
                    if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                elif Ysymbol == '<':
                    if Less(wb, sheet, year, ycell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                elif Ysymbol == '<=':
                    if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
                else:
                    if sheet[ycell].value == year:
                        PrintMovie(wb, sheet, i) # prints the movie
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
    elif year != 'n' and length != 'n' and genre == 'n':
        for i in range(2, sheet.max_row+1):
            lcell = 'D' + str(i)
            ycell = 'B' + str(i)
            if symbol == '<':
                if Less(wb, sheet, length, lcell, i, ct) == True:
                    if Ysymbol == '<':
                        if Less(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '<=':
                        if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>':
                        if Greater(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>=':
                        if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    else:
                        if sheet[ycell].value == year:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
            elif symbol == '<=':
                if LessEqual(wb, sheet, length, lcell, i, ct) == True:
                    if Ysymbol == '<':
                        if Less(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '<=':
                        if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>':
                        if Greater(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>=':
                        if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    else:
                        if sheet[ycell].value == year:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
            elif symbol == '>':
                if Greater(wb, sheet, length, lcell, i, ct) == True:
                    if Ysymbol == '<':
                        if Less(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '<=':
                        if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>':
                        if Greater(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>=':
                        if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    else:
                        if sheet[ycell].value == year:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
            elif symbol == '>=':
                if GreaterEqual(wb, sheet, length, lcell, i, ct) == True:
                    if Ysymbol == '<':
                        if Less(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '<=':
                        if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>':
                        if Greater(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    elif Ysymbol == '>=':
                        if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
                    else:
                        if sheet[ycell].value == year:
                            PrintMovie(wb, sheet, i) # prints the movie
                            ImportMovie(wb, sheet, i, getindex(wb))
                            ct += 1
            else:
                if str(sheet[lcell].value).lower() == length.lower() and str(sheet[ycell].value).lower() == year:
                    PrintMovie(wb, sheet, i)
                    ImportMovie(wb, sheet, i, getindex(wb))
                    ct += 1
    elif year != 'n' and length != 'n' and genre != 'n':
        for i in range(2, sheet.max_row+1):
            lcell = 'D' + str(i)
            ycell = 'B' + str(i)
            gcell = 'C' + str(i)
            if str(sheet[gcell].value).lower() == genre.lower():
                if symbol == '<':
                    if Less(wb, sheet, length, lcell, i, ct) == True:
                        if Ysymbol == '<':
                            if Less(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '<=':
                            if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>':
                            if Greater(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>=':
                            if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        else:
                            if sheet[ycell].value == year:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                elif symbol == '<=':
                    if LessEqual(wb, sheet, length, lcell, i, ct) == True:
                        if Ysymbol == '<':
                            if Less(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '<=':
                            if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>':
                            if Greater(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>=':
                            if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        else:
                            if sheet[ycell].value == year:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                elif symbol == '>':
                    if Greater(wb, sheet, length, lcell, i, ct) == True:
                        if Ysymbol == '<':
                            if Less(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '<=':
                            if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>':
                            if Greater(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>=':
                            if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        else:
                            if sheet[ycell].value == year:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                elif symbol == '>=':
                    if GreaterEqual(wb, sheet, length, lcell, i, ct) == True:
                        if Ysymbol == '<':
                            if Less(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '<=':
                            if LessEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>':
                            if Greater(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        elif Ysymbol == '>=':
                            if GreaterEqual(wb, sheet, year, ycell, i, ct) == True:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                        else:
                            if sheet[ycell].value == year:
                                PrintMovie(wb, sheet, i) # prints the movie
                                ImportMovie(wb, sheet, i, getindex(wb))
                                ct += 1
                else:
                    if str(sheet[lcell].value).lower() == length.lower() and str(sheet[ycell].value).lower() == year:
                        PrintMovie(wb, sheet, i)
                        ImportMovie(wb, sheet, i, getindex(wb))
                        ct += 1
    else:
        PrintErrors('Must use at least 2 filters, Try again ')
        MoreFilters(wb, sheet)

    if ct == 0:
        print()
        PrintErrors('There are none')
    else:
        print()
        PrintStyle('All movies with requested parameters are in the database under \'ParMovies\' ' + '\n' + 'Would you like to look at them? (y/n)')
        print()
        wb.save('moviedb.xlsx')
        look = input('> ')
        if look.lower().startswith('y'):
            try:
                os.system('start SCALC.EXE moviedb.xlsx')
            except:
                try:
                    os.system('start EXCEL.EXE moviedb.clsx')
                except:
                    PrintError('Cannot open up the file')
        else:
            main()
    main()



    
    

def getType(wb, sheet, tname): # get the right format for length
    HMMFormat = re.compile(r'\d H \d\d m')
    HMFormat = re.compile(r'\d H \d m')
    HFormat = re.compile(r'\d H')
    MMFormat = re.compile(r'\d\d m')
    MFormat = re.compile(r'\d m')
    mo = HMMFormat.search(tname)
    if mo == None:
        mo = HMFormat.search(tname)
        if mo == None:
            mo = HFormat.search(tname)
            if mo == None:
                mo = MMFormat.search(tname)
                if mo == None:
                    mo = MFormat.search(tname)
                    if mo == None: # if user typed in a wrong format
                        return None
    if len(tname) == 8:
        hm = HMMFormat
    elif len(tname) == 7:
        hm = HMFormat
    elif len(tname) == 3:
        if tname.endswith('H'):
            hm = HFormat
        elif tname.endswith('m'):
            hm = MFormat
    elif len(tname) == 4:
        hm = MMFormat
    
    return hm

def PrintTypes(wb, sheet, c): # print all the types (genre, year)
    print()
    print('#'*50)
    print()
    types = []
    for i in range(2, sheet.max_row+1): # for every row
        cell = c + str(i)
        if sheet[cell].value not in types: # add it to the list if it is not in it
            types.append(sheet[cell].value)

    if c == 'D': # Length
        lens = [] # array of lengths in hour minute format
        for i in range(2, sheet.max_row+1):
            cell = 'D' + str(i)
            if sheet[cell].value not in lens:
                lens.append(sheet[cell].value)

        # change to minutes
        for i in range(len(lens)-1):
            lens[i] = HtoM(lens[i])

        # change to ints so lens.sort() will work
        for i in range(len(lens)-1):
            lens[i] = int(lens[i])
        
        # deal with the 'None' at the end for some reason, may have to get rid of this at some point
        del lens[len(lens)-1]

        # sort the list
        lens.sort()

        # change back to H M format
        for i in range(len(lens)):
            lens[i] = MtoH(lens[i])
            if len(str(lens[i])) == 2 or len(str(lens[i])) == 1:
                lens[i] = str(lens[i]) + ' m'
        
        types = lens


    for i in range(len(types)-1): # -1 for "None"
        print(types[i])
        print()
    
    
    print()
    print('#'*50)
    
    main()
def HtoM(l): # conver hours to minutes

    if len(l) == 8: # "1 H 10 m"
        h = int(l[0])
        m = l[4:6]
        m = int(m)
        while h > 0:
            m += 60
            h -= 1
    elif len(l) == 7: # " 1 H 5 m"
        h = int(l[0])
        m = l[4]
        m = int(m)
        while h > 0:
            m += 60
            h -= 1
    elif len(l) == 4: # "45 m"
        m = l[:2]
        m = int(m)
    elif len(l) == 3: # "1 H" / "4 M"
        if l.endswith('H'): # "2 H"
            m = int(l[0]) * 60
        elif l.endswith('m'): # "4 M"
            m = int(l[0])

    return m

def MtoH(l): # convert minutes to hours and minutes
    h = 0
    while int(l) >= 60:
        l = int(l) - 60
        h += 1
    if h > 0:
        if l == 0:
            t = '{} H'.format(h) # just hours
        else:
            t = '{} H {} m'.format(h, l) # just minutes
    else:
        t = l
    return t

def Equal(wb, sheet, tname, cell, i, ct, hm): # searches for movies that are equal to the given parameter
    if str(sheet[cell].value).lower() == tname.lower():
        return True
    return False

def Greater(wb, sheet, tname, cell, i, ct): # searches for movies greater than the given parameter
    cellvalue = str(sheet[cell].value)
    if cellvalue == 'N/A' or cellvalue.endswith('m'): # if the type is 'N/A' or if the scraping fucked up
        return False
    tname = tname[2:] # gets rid of the expression symbols
    if cell.startswith('D'): # length
        cellvalue = int(HtoM(cellvalue))
        tname = int(HtoM(tname))
    cellvalue = str(cellvalue)
    if not cellvalue.startswith('N'): # N/A for some movies
        if int(cellvalue) > int(tname):
            return True
    return False

def Less(wb, sheet, tname, cell, i, ct): # searches for movies that are less than the given parameter
    cellvalue = str(sheet[cell].value)
    if cellvalue == 'N/A' or cellvalue.endswith('m'):
        return False
    tname = tname[2:] # gets rid of the expression symbols
    if cell.startswith('D'): # length
        cellvalue = int(HtoM(cellvalue))
        tname = int(HtoM(tname))
    cellvalue = str(cellvalue)
    if not cellvalue.startswith('N'): # N/A for some movies
        if int(cellvalue) < int(tname):
            return True
    return False

def GreaterEqual(wb, sheet, tname, cell, i, ct): # searches for movies that are greater or equal to the given parameter
    cellvalue = str(sheet[cell].value)
    if cellvalue == 'N/A' or cellvalue.endswith('m'):
        return False
    tname = tname[3:] # gets rid of the expression symbols
    if cell.startswith('D'): # length
        cellvalue = int(HtoM(cellvalue))
        tname = int(HtoM(tname))
    cellvalue = str(cellvalue)
    if not cellvalue.startswith('N'): # N/A for some movies
        if int(cellvalue) >= int(tname):
            return True
    return False

def LessEqual(wb, sheet, tname, cell, i, ct): # searches for movies that are less or equal to the given parameter
    cellvalue = str(sheet[cell].value)
    if cellvalue == 'N/A' or cellvalue.endswith('m'):
        return False
    tname = tname[3:] # gets rid of the expression symbols
    if cell.startswith('D'): # length
        cellvalue = int(HtoM(cellvalue))
        tname = int(HtoM(tname))
    cellvalue = str(cellvalue)
    if not cellvalue.startswith('N'): # N/A for some movies
        if int(cellvalue) <= int(tname):
            return True
    return False

def Search(wb, sheet): # search by text
    print()
    inlist = False
    PrintStyle('Enter the movie you want to search ')
    print()
    movie = input('> ')
    print()
    movie = movie.lower() # changes to all lowercase to avoid confusion
    for i in range(2, sheet.max_row+1):
        cell = 'A' + str(i)
        if movie in str(sheet[cell].value).lower(): # searches if the movie searched is in the current movie we are looking at
            PrintMovie(wb, sheet, i) # prints the movie
            ImportMovie(wb, sheet, i, getindex(wb))
            inlist = True
    if inlist == False: # if movie is not in database
        PrintErrors('\"{}\" is nowehere in the database!'.format(movie))
    else:
        print()
        PrintStyle('All movies with requested parameters are in the database under \'ParMovies\' ' + '\n' + 'Would you like to look at them? (y/n)')
        print()
        wb.save('moviedb.xlsx')
        look = input('> ')
        if look.lower().startswith('y'):
            os.system('start SCALC.EXE moviedb.xlsx')
        else:
            main()
    main()

def Random(wb, sheet): # gets a random movie from the database
    print()
    PrintStyle('How would you like to get a random movie?')
    ans = input('''

    (r)    Random                 (g)    Genre

    (y)    Year                   (l)    Length 

    (m)    Main Menu
    
> ''')
    if ans.lower().startswith('r'): # true random movie
        rand = random.randint(2,10)
        PrintMovie(wb, sheet, rand)
        link = sheet['E' + str(rand)].value
        pyperclip.copy(link)
        PrintStyle('The link for the movie has been added to your clipboard')
    elif ans.lower().startswith('g'): # random from specific genre
        RandomType(wb, sheet, 'genre', 'C')
    elif ans.lower().startswith('y'): # random from specific year
        RandomType(wb, sheet, 'year', 'B')
    elif ans.lower().startswith('l'): # random by specific length
        RandomType(wb, sheet, 'length', 'D')
    elif ans.lower().startswith('m'):
        main()
    else: # not the right command
        print()
        PrintErrors('That didn\'t work, Try again')
        Random(wb, sheet)
    main()

def RandomType(wb, sheet, t, c): # gets random movie based on parameter
    print()
    ct = []
    PrintStyle('Enter the {} you want'.format(t))
    print()
    r = input('> ')
    for i in range(2, sheet.max_row+1): # looks through every movie in the database
        cell = c + str(i)
        if cell == 'N/A':
            continue
        if str(sheet[cell].value).lower() == r.lower(): # finds the right movie
            ct.append(i) # add to the list of all movies with the correct parameters
    if len(ct) == 0: # if there are no movies with the given parameter
        PrintErrors('There are none ')
        main()
    else: # otherwise grab a random movie from the list and print it
        rand = random.randint(0,len(ct)-1)
        ri = ct[rand]
        PrintMovie(wb, sheet, ri) # print the movie
        link = sheet['E' + str(ri)].value
        pyperclip.copy(link)
        PrintStyle('The link for the movie has been added to your clipboard')


def PrintMovie(wb, sheet, i): # prints the movie
    print()
    print('#'*50)
    print('''
Name:    {}
Year:    {}
Genre:   {}
Length:  {}
Link:    {}
    '''.format(sheet['A' + str(i)].value, 
    sheet['B' + str(i)].value, 
    sheet['C' + str(i)].value, 
    sheet['D' + str(i)].value, 
    sheet['E' + str(i)].value))

    print()
    print('#' * 50)
    print()

def ImportMovie(wb, sheet, i, m): # get the specific movies into the sheet
    ParSheet = wb['ParMovies']
    ParSheet.cell(row = m, column = column_index_from_string('A')).value = sheet['A' + str(i)].value
    ParSheet.cell(row = m, column = column_index_from_string('B')).value = sheet['B' + str(i)].value
    ParSheet.cell(row = m, column = column_index_from_string('C')).value = sheet['C' + str(i)].value
    ParSheet.cell(row = m, column = column_index_from_string('D')).value = sheet['D' + str(i)].value
    ParSheet.cell(row = m, column = column_index_from_string('E')).value = sheet['E' + str(i)].value
    for j in range(2, ParSheet.max_row+1): # change the row height
        ParSheet.row_dimensions[j].height = 40
    wb.save('moviedb.xlsx')

def getindex(wb): # get the index where the next movie is to be placed
    ParSheet = wb['ParMovies']
    for i in range(2, ParSheet.max_row+1): 
        if ParSheet['A' + str(i)].value == ' ': # if this cell is empty, return this cell
            return i
    return ParSheet.max_row+1

def ClearSheet(wb): # clear the parameter sheet
    ParSheet = wb['ParMovies']

    ParSheet = wb['ParMovies']
    for i in range(2, ParSheet.max_row): # change each cell to nothing
        ParSheet['A' + str(i)] = ' '
        ParSheet['B' + str(i)] = ' '
        ParSheet['C' + str(i)] = ' '
        ParSheet['D' + str(i)] = ' '
        ParSheet['E' + str(i)] = ' '

def PrintStyle(t): # printing questions for user
    print()
    print()
    print('`'*50)
    print()
    print(t)
    print()
    print('`'*50)

def PrintErrors(t): # printing errors
    print()
    print('?'*50)
    print()
    print(t)
    print()
    print('?'*50)


if __name__ == '__main__': # run main
    main()