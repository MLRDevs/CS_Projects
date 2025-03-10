# Alex Miller
# Working With Excel Spreadsheets

# Reading Excel Documents

import openpyxl
wb = openpyxl.load_workbook('example.xlsx') # loads the excel document
print(type(wb)) # <class 'openpyxl.workbook.workbook.Workbook'>

# Getting Sheets from the Workbook

print(wb.sheetnames) # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3']
print(sheet) # <Worksheet "Sheet3">
print(type(sheet)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(sheet.title) # Sheet3
anothersheet = wb.active
print(anothersheet) # <Worksheet "Sheet1">

print()

# Getting Cells from the Sheets

sheet = wb['Sheet1'] # getting a sheet from the workbook
print(sheet['A1']) # getting a cell from the sheet
print(sheet['A1'].value) # get the value from the cell
c = sheet['B1'] # get another cell from the sheet
print(c.value)

# get row, column, and value from cell
print('Row %s, Column %s, is %s' % (c.row, c.column, c.value))
print(sheet['C1'].value)

print()

print(sheet.cell(row=1, column=2)) # <Cell 'Sheet1'.B1>
print(sheet.cell(row=1, column=2).value) # 'Apples'
print()

for i in range(1, 8, 2): # go through every other row
    print(i, sheet.cell(row=i, column=2).value)

print()

print(sheet.max_row) # highest row number
print(sheet.max_column) # highest column number

print()

# Converting Between Column Letters and Numbers

from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1)) # A
print(get_column_letter(2)) # B
print(get_column_letter(27)) # AA
print(get_column_letter(900)) # AHP
print(get_column_letter(sheet.max_column)) # C
print(column_index_from_string('A')) # 1
print(column_index_from_string('AA')) # 27

print()

# Getting Rows and Columns from the Sheets

print(tuple(sheet['A1':'C3'])) # get all cells from A1 to C3
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

print()

print(list(sheet.columns)[1]) # get second column's cells
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)

print()

# Workbooks, Sheets, Cells
# reading a cell review

# 1, import openpyxl
# 2, call the openpyxl.load_workbook()
# 3, get a Workbook object
# 4, use the active or sheetnames attribute
# 5, get a Worksheet object
# 6, use indexing or cell() sheet method with row and column keyword arguments
# 7, get a cell object
# 8, read the cell object's value attribute

# Creating and Saving Excel Documents

wb = openpyxl.Workbook() # creat a blank workbook
print(wb.sheetnames) # ['Sheet']
sheet = wb.active
print(sheet.title) # 'Sheet'
sheet.title = 'Spam Bacon Eggs Sheet' # changes the title
print(sheet.title) # ['Spam Bacon Eggs Sheet']

wb.save('example_copy.xlsx') # save the workbook

print()

# Creating and Removing Sheets

wb = openpyxl.Workbook()
wb.create_sheet() # adds a new sheet
print(wb.sheetnames) # ['Sheet', 'Sheet1']
wb.create_sheet(index=0, title='First Sheet') # creates a new sheet at index 0
print(wb.sheetnames) # ['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet') # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
del wb['Middle Sheet'] # delete sheets
del wb['Sheet1']
print(wb.sheetnames) # ['First Sheet', 'Sheet']

print()

# Writing Values to Cells

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, World!' # edit the cells value
print(sheet['A1'].value)

print()

# Setting the Font Style of Cells

from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # create a font
sheet['A1'].font = italic24Font # apply font to A1
sheet['A1'] = 'Hello, World!'
wb.save('styles.xlsx')

print()

# Font Objects

# Keyword Argument == Data Type == Description

# name == String == The font name, ex: calibri or Times New Roman
# size == Integer == point size
# bold == Boolean == True for bold font
# italic == Boolean == True, for italic font

wb = openpyxl.load_workbook('styles.xlsx')
sheet = wb['Sheet']

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A5'].font = fontObj1
sheet['A5'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('styles.xlsx')

print()

# Formulas

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # set the formula
wb.save('writeFormula.xlsx')

print()

# Setting Row Height and Column Width

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall Row'
sheet['B2'] = 'Wide Column'

sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

print()

# Merging and Unmerging Cells

wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # merge all these cells
sheet['A1'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5') # merge these two cells
sheet['C5'] = 'Two merged cells'
wb.save('merged.xlsx')

wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3') # unmerge the cells
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

print()

# Freezing Panes

# sheet.freeze_panes = 'A2' ==> Row 1
# sheet.freeze_panes = 'B1' ==> Column A
# sheet.freeze_panes = 'C1' ==> Columns A and B
# sheet.freeze_panes = 'C2' ==> Row 1 and Columns A and B
# sheet.freeze_panes = 'A1' ==> no frozen panes
# sheet.freeze_panes = None ==> no frozen panes

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2' # freeze the rows above A2
wb.save('freezeExample.xlsx')

print()

# Charts

# 1, create a reference object from a rectangular selection of cells
# 2, create a series object ny passing in the reference object
# 3, create a chart object
# 4, append the series object to the chart object
# 5, add the char object to the worksheet object, optionally specifying which cell should be the top-left corner of the chart

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')






