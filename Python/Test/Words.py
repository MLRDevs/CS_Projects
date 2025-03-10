# Alex Miller
# search excel sheet for specific word

import openpyxl, sys, random

def main():
    '''
    name = input('Enter the name of the file: ')
    try:
        wb = openpyxl.load_workbook(name)
    except:
        print('That didn\'t work, Try again')
        print()
        main()
    '''
    print()
    wb = openpyxl.load_workbook('words.xlsx') # open the workbook
    sheet = wb['Sheet1']
    command = input('''
Would you like to:

    (s)     search
    (a)     add
    (e)     edit
    (cls)   clear screen
    (ls)    list size
    (scp)   sort, copy paste
    (q)     quit

> ''')
    cmd(wb, sheet, command)

def cmd(wb, sheet, command): # process command
    if command.lower() == 's': # search
        word = input('Enter the word you want to search (m to main menu): ')
        if word == 'm':
            main()
        Search(wb, sheet, word)
    elif command.lower().startswith('a'): # add
        Add(wb, sheet)
    elif command.lower().startswith('e'): # edit
        Edit(wb, sheet)
    elif command.lower().startswith('c'): # clear
        ClearScreen()
    elif command.lower().startswith('l'): # list size
        print()
        print('The total size of the list is: ' + str(sheet.max_row))
        main()
    elif command.lower().startswith('scp'): # sort then copy paste
        Sort(wb, sheet)
    elif command.lower().startswith('q'): # quit
        print()
        print('Thanks!')
        sys.exit()
    else: # dumbass
        print()
        print('That didn\'t work, Try again')
        main()

def Search(wb, sheet, word): # searching the list
    print()
    sheet = wb['Sheet1']
    for i in range(1, sheet.max_row+1): # go through every cell
        cell = 'A' + str(i)
        if sheet[cell].value == word or str(sheet[cell].value).lower() == word.lower():
            print()
            print(word + ' is in the list at cell ' + cell) # word is in the list
            break
        elif i == sheet.max_row and sheet[cell].value != word: # word is not in the list
            print()
            print(word + ' is not in the list')
            print()
            add = input('Would you like to add it? (y/n): ') # word is not in list, ask if user wants to add it
            if add.lower().startswith('y'):
                AddIt(wb, sheet, word)
            else:
                main()

    print()      
    main()

def Add(wb, sheet): # adding to the list
    cell = 'A' + str(sheet.max_row+1)
    print()
    word = input('Enter the word you want to add (m to main menu): ')
    if word.lower().startswith('m'): # back to the main menu
        main()
    sheet[cell] = word # add the word to the next empty cell
    print()
    print(word + ' has been added at cell ' + cell)
    wb.save('words.xlsx') # save the file
    main()

def Edit (wb, sheet): # edit an existing word in the list
    print()
    cellword = ''
    inlist = False
    word = input('Enter the word you want to edit: ') # word to edit
    nw = input('Enter what you want to add instead: ') # new word to replace
    if Check(wb, sheet, nw) == True: # check if the new word is already in the list
        print()
        print(nw + ' is already in the list')
    else: # if the new word is not in the list
        cell = ''
        for j in range(1, sheet.max_row+1): # find the cell with the old word and replace it
            cell = 'A' + str(j)
            if sheet[cell].value == word or str(sheet[cell].value).lower() == word.lower():
                sheet[cell] = nw
        print()
        print(nw + ' has been added.')
        wb.save('words.xlsx') # save the file
        main()

    if Check(wb, sheet, nw) == True: # ask if the user wants to edit a different word if the new word is already in the list
        print()
        again = input('Would you like to edit a new word? (y/n): ')
        if again.lower().startswith('y'):
            Edit(wb, sheet)
        else:
            main()


def ClearScreen(): # clear the screen
    for i in range(50):
        print()
    main()

def AddIt(wb, sheet, word): # add to the list after searching and finding out it's not in the list
    cell = 'A' + str(sheet.max_row+1)
    print()
    sheet[cell] = word # add the word
    print()
    print(word + ' has been added to cell ' + cell)
    wb.save('words.xlsx') # save the file
    main()
        
def Sort(wb, sheet): # shuffling the list and pasting it in column C
    print()
    words = [] # list of words
    for i in range(1,sheet.max_row+1): # add all the words to the list
        cell = 'A' + str(i)
        words.append(sheet[cell].value)

    rand = random.randint(50,100) # random value fro 50 to 100
    for i in range(rand): # shuffle the list that many times
        random.shuffle(words)
    for i in range(len(words)): # go down each cell in column C and paste all the words in the shuffled list
        cell = 'C'
        cell = cell + str(i+1)
        sheet[cell].value = words[i]
    
    print('The list has been sorted and moved.')
    wb.save('words.xlsx') # save the file
    print()
    main()

def Check(wb, sheet, word): # check whether a word is in the list or not
    for i in range(1, sheet.max_row+1):
        cell = 'A' + str(i)
        if sheet[cell].value == word or str(sheet[cell].value).lower() == word.lower(): # if the word is in the current cell
            return True
    return False

if __name__ == '__main__':
    main()